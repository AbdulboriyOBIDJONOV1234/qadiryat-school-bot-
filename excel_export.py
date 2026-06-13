from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    "№",
    "F.I.Sh. (to'liq ism)",
    "Tug'ilgan sana",
    "Sinf",
    "Manzil",
    "Telefon raqam",
    "Ro'yxatdan o'tgan sana",
]

COLUMN_WIDTHS = [5, 32, 16, 10, 24, 18, 20]


def build_excel(rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ro'yxat"

    sheet.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for index, row in enumerate(rows, start=1):
        _id, full_name, birth_date, grade, location, phone, created_at = row
        sheet.append([index, full_name, birth_date, f"{grade}-sinf", location, phone, created_at])

    for col_index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
